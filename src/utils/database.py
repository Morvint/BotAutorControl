import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

class Database():
    def __init__(self, db_name):
        self.connection = sqlite3.connect(db_name)
        self.cursor = self.connection.cursor()
        self.create_db()

    def create_db(self):
        try:
            query = ('CREATE TABLE IF NOT EXISTS `users`('
                     '`id` INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,'
                     '`user_name` TEXT,'
                     '`user_phone` TEXT,'
                     '`telegram_id` TEXT);')
            self.cursor.execute(query)
            self.connection.commit()
            query = ('CREATE TABLE IF NOT EXISTS `wait_reg`('
                     '`id` INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,'
                     '`user_name` TEXT,'
                     '`user_phone` TEXT,'
                     '`telegram_id` TEXT);')
            self.cursor.execute(query)
            self.connection.commit()
            query = ('CREATE TABLE IF NOT EXISTS `questions`('
                     '`id` INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,'
                     '`block` TEXT,'
                     '`korpus` TEXT,'
                     '`company` TEXT,'
                     '`fullname` TEXT,'
                     '`paragraph_and_numlist` TEXT,'
                     '`question` TEXT,'
                     '`photo` TEXT,'
                     '`date_question` TEXT,'
                     '`date_result` TEXT,'
                     '`status` TEXT,'
                     '`result` TEXT);')
            self.cursor.execute(query)
            self.connection.commit()
        except sqlite3.Error as Error:
            print('Ошибка при создании: ', Error)

    def add_user(self, user_name, user_phone, telegram_id):
        self.cursor.execute(f'INSERT INTO `users`(user_name, user_phone, telegram_id) VALUES (?, ?, ?)', (user_name, user_phone, telegram_id))
        self.connection.commit()

    def add_user_wait_reg(self, user_name, user_phone, telegram_id):
        self.cursor.execute(f'INSERT INTO `wait_reg`(user_name, user_phone, telegram_id) VALUES (?, ?, ?)', (user_name, user_phone, telegram_id))
        self.connection.commit()

    def add_question_no_photo(self, block, korpus, fullname, paragraph, numlist, question, photo, date_question):
        paragraph_and_numlist = paragraph + "/" + numlist
        self.cursor.execute(f'INSERT INTO `questions`(block, korpus, fullname, paragraph_and_numlist, question, photo, date_question) VALUES (?, ?, ?, ?, ?, ?, ?)', (block, korpus, fullname, paragraph_and_numlist, question, photo, date_question))
        self.connection.commit()

    def select_user_id(self, telegram_id):
        users = self.cursor.execute('SELECT * FROM `users` WHERE telegram_id = ?', (telegram_id,))
        return users.fetchone()
    
    def select_wait_reg_user_id(self, telegram_id):
        users = self.cursor.execute('SELECT * FROM `wait_reg` WHERE telegram_id = ?', (telegram_id,))
        return users.fetchone()
    
    def count_wait_reg(self):
        count = self.cursor.execute('SELECT COUNT(*) FROM `wait_reg`')
        return count.fetchone()[0]
    
    def first_wait_reg_user(self):
        user = self.cursor.execute('SELECT * FROM `wait_reg`')
        return user.fetchone()
    
    def get_questions_table(self):
        query = 'SELECT * FROM `questions`'
        data = pd.read_sql_query(query, self.connection)
        data.columns = ['№', 'Квартал', 'Корпус', 'Компания', 'ФИО', 'Название раздела/Номер листа', 'Вопрос', 'Фото', 'Дата постановки', 'Дата выполнения', 'Статус', 'Результат']
        data.to_excel('table.xlsx', index=False)
        
        workbook = load_workbook('table.xlsx')
        worksheet = workbook['Sheet1']
        for col in range(1, len(data.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:  # Не все ячейки могут содержать строковые значения
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2)  # Добавляем немного места
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save('table.xlsx')
    
    def del_user_wait_reg(self, telegram_id):
        self.cursor.execute(f'DELETE FROM `wait_reg` WHERE telegram_id = ?', (telegram_id,))
        self.connection.commit()
        
    def __del__(self):
        self.cursor.close()
        self.connection.close()